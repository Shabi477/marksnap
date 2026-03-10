from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


# Colors
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
KEY_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
SCORE_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def generate_results_excel(test_name: str, student_results: list, answer_keys: list) -> bytes:
    """
    Generate an Excel file with question-level results.
    Layout: Students as columns, questions as rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Question Analysis"

    # --- Build the sheet ---
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(student_results) + 3)
    title_cell = ws.cell(row=1, column=1, value=f"MarkSnap — {test_name} — Question Level Analysis")
    title_cell.font = Font(bold=True, size=16, color="6366F1")
    title_cell.alignment = Alignment(horizontal="center")

    # Row 3: Headers
    header_row = 3
    headers = ["Question", "Section", "Answer"]  # First 3 columns

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Student name headers
    for s_idx, student in enumerate(student_results):
        col = s_idx + 4
        cell = ws.cell(row=header_row, column=col, value=student.student_name)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 5

    # Row 4: Class info
    class_row = header_row + 1
    ws.cell(row=class_row, column=1, value="Class").font = BOLD_FONT
    ws.cell(row=class_row, column=2, value="").font = BOLD_FONT
    ws.cell(row=class_row, column=3, value="").font = BOLD_FONT
    for s_idx, student in enumerate(student_results):
        col = s_idx + 4
        cell = ws.cell(row=class_row, column=col, value=student.class_name)
        cell.font = Font(size=8, color="6B7280")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Row 5: Student ID
    id_row = class_row + 1
    ws.cell(row=id_row, column=1, value="ID").font = BOLD_FONT
    for s_idx, student in enumerate(student_results):
        col = s_idx + 4
        cell = ws.cell(row=id_row, column=col, value=student.student_code)
        cell.font = Font(size=8, color="6B7280")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Questions data rows
    data_start_row = id_row + 1

    for q_idx, (q_num, section_name, correct_answer) in enumerate(answer_keys):
        row = data_start_row + q_idx
        q_key = f"Q{q_num}"

        # Question number
        cell = ws.cell(row=row, column=1, value=q_num)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Section
        cell = ws.cell(row=row, column=2, value=section_name)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Correct answer
        cell = ws.cell(row=row, column=3, value=correct_answer)
        cell.fill = KEY_FILL
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Student answers
        for s_idx, student in enumerate(student_results):
            col = s_idx + 4
            answer = student.answers.get(q_key, "")
            is_correct = student.correct.get(q_key, False)

            cell = ws.cell(row=row, column=col, value=answer or "-")
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            cell.font = NORMAL_FONT

            if answer:
                cell.fill = GREEN_FILL if is_correct else RED_FILL

    # Summary rows
    summary_start = data_start_row + len(answer_keys) + 1

    # Score row
    ws.cell(row=summary_start, column=1, value="Score").font = SCORE_FONT
    ws.cell(row=summary_start, column=1).border = THIN_BORDER
    for s_idx, student in enumerate(student_results):
        col = s_idx + 4
        cell = ws.cell(row=summary_start, column=col, value=f"{student.score}/{student.total}")
        cell.font = SCORE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Percentage row
    ws.cell(row=summary_start + 1, column=1, value="Percentage").font = BOLD_FONT
    ws.cell(row=summary_start + 1, column=1).border = THIN_BORDER
    for s_idx, student in enumerate(student_results):
        col = s_idx + 4
        cell = ws.cell(row=summary_start + 1, column=col, value=f"{student.percentage}%")
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # --- Question difficulty row (% of students who got it right) ---
    diff_row = summary_start + 3
    ws.cell(row=diff_row, column=1, value="Item Difficulty (% Correct)").font = BOLD_FONT

    for q_idx, (q_num, section_name, correct_answer) in enumerate(answer_keys):
        q_key = f"Q{q_num}"
        correct_count = sum(1 for s in student_results if s.correct.get(q_key, False))
        total_students = len(student_results) or 1
        difficulty = round(correct_count / total_students * 100, 1)

        row = diff_row + 1 + q_idx
        ws.cell(row=row, column=1, value=q_num).border = THIN_BORDER
        ws.cell(row=row, column=2, value=section_name).border = THIN_BORDER

        cell = ws.cell(row=row, column=3, value=f"{difficulty}%")
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

        # Color code difficulty
        if difficulty >= 80:
            cell.fill = GREEN_FILL
        elif difficulty < 50:
            cell.fill = RED_FILL

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10

    # Freeze panes so headers stay visible
    ws.freeze_panes = "D7"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
