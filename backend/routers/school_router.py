from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from database import get_db
from models import Teacher, School, ClassGroup, Student, teacher_classes, Test, TestAssignment
from schemas import (
    SchoolResponse, ClassCreate, ClassResponse, TeacherResponse,
    AssignTeacher, StudentResponse, StudentTransfer,
    TestAssignmentCreate, TestAssignmentResponse,
)
from auth import get_current_teacher
from datetime import datetime
import uuid
import io

router = APIRouter(prefix="/api/school", tags=["school"])


def require_hod(teacher: Teacher):
    if teacher.role != "hod":
        raise HTTPException(status_code=403, detail="Only HOD can perform this action")


def _class_response(c: ClassGroup, db: Session) -> ClassResponse:
    teacher_names = [t.name for t in c.teachers]
    return ClassResponse(
        id=c.id, name=c.name, academic_year=c.academic_year,
        key_stage=c.key_stage,
        school_id=c.school_id, owner_id=c.owner_id,
        student_count=len(c.students), teacher_names=teacher_names,
        created_at=c.created_at,
    )


@router.get("/", response_model=SchoolResponse)
def get_school(
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    if not teacher.school_id:
        raise HTTPException(status_code=404, detail="Not part of a school")
    return teacher.school


@router.get("/invite-code")
def get_invite_code(
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    return {"invite_code": teacher.school.invite_code}


@router.post("/regenerate-invite")
def regenerate_invite_code(
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    from models import _generate_invite_code
    teacher.school.invite_code = _generate_invite_code()
    db.commit()
    return {"invite_code": teacher.school.invite_code}


# --- Teachers ---
@router.get("/teachers", response_model=list[TeacherResponse])
def list_teachers(
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    teachers = db.query(Teacher).filter(Teacher.school_id == teacher.school_id).all()
    result = []
    for t in teachers:
        result.append({
            "id": t.id, "email": t.email, "name": t.name,
            "role": t.role, "tier": t.tier, "school_id": t.school_id,
            "school_name": t.school.name if t.school else None,
            "school_type": t.school.school_type if t.school else None,
            "region": t.school.region if t.school else None,
            "school_tier": t.school.tier if t.school else None,
            "created_at": t.created_at,
        })
    return result


@router.put("/teachers/{teacher_id}/role")
def update_teacher_role(
    teacher_id: int,
    data: dict,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    """Update a teacher's role (HOD/school_admin only)."""
    require_hod(teacher)
    if teacher_id == teacher.id:
        raise HTTPException(status_code=403, detail="Cannot modify your own role")
    target = db.query(Teacher).filter(
        Teacher.id == teacher_id,
        Teacher.school_id == teacher.school_id,
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Teacher not found")
    new_role = data.get("role")
    if new_role not in ("teacher", "sendco", "hod", "school_admin"):
        raise HTTPException(status_code=400, detail="Invalid role")
    target.role = new_role
    db.commit()
    return {"id": target.id, "name": target.name, "role": target.role}


# --- Classes (HOD creates for school) ---
@router.get("/classes", response_model=list[ClassResponse])
def list_school_classes(
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    classes = db.query(ClassGroup).filter(ClassGroup.school_id == teacher.school_id).all()
    return [_class_response(c, db) for c in classes]


@router.post("/classes", response_model=ClassResponse)
def create_school_class(
    data: ClassCreate,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    class_group = ClassGroup(
        name=data.name, academic_year=data.academic_year,
        key_stage=data.key_stage,
        school_id=teacher.school_id, owner_id=teacher.id,
    )
    db.add(class_group)
    db.commit()
    db.refresh(class_group)
    return _class_response(class_group, db)


@router.delete("/classes/{class_id}")
def delete_school_class(
    class_id: int,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    class_group = db.query(ClassGroup).filter(
        ClassGroup.id == class_id, ClassGroup.school_id == teacher.school_id
    ).first()
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")
    db.delete(class_group)
    db.commit()
    return {"message": "Class deleted"}


# --- Teacher-Class Assignment ---
@router.post("/assign-teacher")
def assign_teacher_to_class(
    data: AssignTeacher,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    target_teacher = db.query(Teacher).filter(
        Teacher.id == data.teacher_id, Teacher.school_id == teacher.school_id
    ).first()
    if not target_teacher:
        raise HTTPException(status_code=404, detail="Teacher not found in your school")

    class_group = db.query(ClassGroup).filter(
        ClassGroup.id == data.class_id, ClassGroup.school_id == teacher.school_id
    ).first()
    if not class_group:
        raise HTTPException(status_code=404, detail="Class not found")

    if target_teacher in class_group.teachers:
        raise HTTPException(status_code=400, detail="Teacher already assigned to this class")

    class_group.teachers.append(target_teacher)
    db.commit()
    return {"message": f"{target_teacher.name} assigned to {class_group.name}"}


@router.delete("/unassign-teacher")
def unassign_teacher_from_class(
    data: AssignTeacher,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    target_teacher = db.query(Teacher).filter(
        Teacher.id == data.teacher_id, Teacher.school_id == teacher.school_id
    ).first()
    class_group = db.query(ClassGroup).filter(
        ClassGroup.id == data.class_id, ClassGroup.school_id == teacher.school_id
    ).first()
    if not target_teacher or not class_group:
        raise HTTPException(status_code=404, detail="Not found")

    if target_teacher in class_group.teachers:
        class_group.teachers.remove(target_teacher)
        db.commit()
    return {"message": "Teacher unassigned"}


# --- Excel Import ---
@router.post("/import-classes")
def import_classes_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    import openpyxl

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Find headers
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    class_col = next((i for i, h in enumerate(headers) if "class" in h), None)
    name_col = next((i for i, h in enumerate(headers) if "name" in h or "student" in h), None)
    id_col = next((i for i, h in enumerate(headers) if "id" in h or "code" in h), None)

    if class_col is None or name_col is None:
        raise HTTPException(status_code=400, detail="Excel must have 'Class' and 'Student Name' columns")

    classes_created = 0
    students_added = 0
    class_cache = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        class_name = str(row[class_col].value or "").strip()
        student_name = str(row[name_col].value or "").strip()
        student_code = str(row[id_col].value or "").strip() if id_col is not None else ""

        if not class_name or not student_name:
            continue

        # Get or create class
        if class_name not in class_cache:
            existing_class = db.query(ClassGroup).filter(
                ClassGroup.school_id == teacher.school_id,
                ClassGroup.name == class_name,
            ).first()
            if existing_class:
                class_cache[class_name] = existing_class
            else:
                new_class = ClassGroup(
                    name=class_name,
                    academic_year=str(datetime.now().year),
                    school_id=teacher.school_id,
                    owner_id=teacher.id,
                )
                db.add(new_class)
                db.flush()
                class_cache[class_name] = new_class
                classes_created += 1

        class_group = class_cache[class_name]

        # Create student
        if not student_code:
            student_code = f"S{uuid.uuid4().hex[:8].upper()}"

        existing_student = db.query(Student).filter(Student.student_code == student_code).first()
        if existing_student:
            continue

        student = Student(name=student_name, student_code=student_code, class_id=class_group.id)
        db.add(student)
        students_added += 1

    db.commit()
    return {
        "message": f"Imported {classes_created} classes and {students_added} students",
        "classes_created": classes_created,
        "students_added": students_added,
    }


# --- Student Search & Transfer ---
@router.get("/students/search", response_model=list[StudentResponse])
def search_students(
    q: str = "",
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    """Search students across all school classes."""
    if not teacher.school_id:
        raise HTTPException(status_code=400, detail="Not part of a school")

    school_class_ids = [
        c.id for c in db.query(ClassGroup).filter(
            ClassGroup.school_id == teacher.school_id
        ).all()
    ]
    if not school_class_ids:
        return []

    query = db.query(Student).filter(Student.class_id.in_(school_class_ids))
    if q:
        query = query.filter(
            (Student.name.ilike(f"%{q}%")) | (Student.student_code.ilike(f"%{q}%"))
        )
    students = query.limit(50).all()

    return [
        StudentResponse(
            id=s.id, name=s.name, student_code=s.student_code,
            class_number=s.class_number, class_id=s.class_id, class_name=s.class_group.name if s.class_group else "",
        )
        for s in students
    ]


@router.post("/students/transfer")
def transfer_student(
    data: StudentTransfer,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    """Move a student to a different class within the school."""
    if not teacher.school_id:
        raise HTTPException(status_code=400, detail="Not part of a school")

    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    # Verify both classes belong to this school
    from_class = db.query(ClassGroup).filter(ClassGroup.id == student.class_id).first()
    to_class = db.query(ClassGroup).filter(ClassGroup.id == data.to_class_id).first()

    if not from_class or from_class.school_id != teacher.school_id:
        raise HTTPException(status_code=403, detail="Student's current class not in your school")
    if not to_class or to_class.school_id != teacher.school_id:
        raise HTTPException(status_code=403, detail="Target class not in your school")

    student.class_id = data.to_class_id
    db.commit()
    return {"message": f"{student.name} moved to {to_class.name}"}


# --- Year Groups ---
@router.get("/year-groups")
def list_year_groups(
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    years = (
        db.query(ClassGroup.academic_year)
        .filter(ClassGroup.school_id == teacher.school_id)
        .distinct()
        .all()
    )
    return sorted([y[0] for y in years])


# --- Test Assignments ---
@router.post("/push-test", response_model=list[TestAssignmentResponse])
def push_test(
    data: TestAssignmentCreate,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)

    test = db.query(Test).filter(Test.id == data.test_id).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    # Verify test belongs to a teacher in this school
    test_owner = db.query(Teacher).filter(Teacher.id == test.teacher_id).first()
    if not test_owner or test_owner.school_id != teacher.school_id:
        raise HTTPException(status_code=403, detail="Test does not belong to your school")

    if not data.class_ids and not data.teacher_ids and not data.year_groups:
        raise HTTPException(status_code=400, detail="Must specify at least one target")

    assignments = []

    # Assign to specific classes
    for class_id in data.class_ids:
        cls = db.query(ClassGroup).filter(
            ClassGroup.id == class_id, ClassGroup.school_id == teacher.school_id
        ).first()
        if not cls:
            continue
        # Skip if already assigned
        existing = db.query(TestAssignment).filter(
            TestAssignment.test_id == data.test_id,
            TestAssignment.class_id == class_id,
        ).first()
        if existing:
            continue
        a = TestAssignment(
            test_id=data.test_id, school_id=teacher.school_id,
            assigned_by=teacher.id, class_id=class_id,
        )
        db.add(a)
        assignments.append(a)

    # Assign to specific teachers
    for tid in data.teacher_ids:
        t = db.query(Teacher).filter(
            Teacher.id == tid, Teacher.school_id == teacher.school_id
        ).first()
        if not t:
            continue
        existing = db.query(TestAssignment).filter(
            TestAssignment.test_id == data.test_id,
            TestAssignment.teacher_id == tid,
        ).first()
        if existing:
            continue
        a = TestAssignment(
            test_id=data.test_id, school_id=teacher.school_id,
            assigned_by=teacher.id, teacher_id=tid,
        )
        db.add(a)
        assignments.append(a)

    # Assign to year groups
    for yg in data.year_groups:
        existing = db.query(TestAssignment).filter(
            TestAssignment.test_id == data.test_id,
            TestAssignment.year_group == yg,
        ).first()
        if existing:
            continue
        a = TestAssignment(
            test_id=data.test_id, school_id=teacher.school_id,
            assigned_by=teacher.id, year_group=yg,
        )
        db.add(a)
        assignments.append(a)

    db.commit()
    for a in assignments:
        db.refresh(a)

    return [_assignment_response(a, db) for a in assignments]


def _assignment_response(a: TestAssignment, db: Session) -> TestAssignmentResponse:
    return TestAssignmentResponse(
        id=a.id, test_id=a.test_id,
        test_name=a.test.name if a.test else "",
        class_id=a.class_id,
        class_name=a.class_group.name if a.class_group else None,
        teacher_id=a.teacher_id,
        teacher_name=a.target_teacher.name if a.target_teacher else None,
        year_group=a.year_group,
        assigned_by_name=a.assigner.name if a.assigner else "",
        created_at=a.created_at,
    )


@router.get("/test-assignments", response_model=list[TestAssignmentResponse])
def list_test_assignments(
    test_id: int = None,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    query = db.query(TestAssignment).filter(TestAssignment.school_id == teacher.school_id)
    if test_id:
        query = query.filter(TestAssignment.test_id == test_id)
    assignments = query.order_by(TestAssignment.created_at.desc()).all()
    return [_assignment_response(a, db) for a in assignments]


@router.delete("/test-assignments/{assignment_id}")
def delete_test_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    teacher: Teacher = Depends(get_current_teacher),
):
    require_hod(teacher)
    a = db.query(TestAssignment).filter(
        TestAssignment.id == assignment_id,
        TestAssignment.school_id == teacher.school_id,
    ).first()
    if not a:
        raise HTTPException(status_code=404, detail="Assignment not found")
    db.delete(a)
    db.commit()
    return {"message": "Assignment removed"}
